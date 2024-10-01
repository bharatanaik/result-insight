import base64
import io
import json
import time
from channels.generic.websocket import WebsocketConsumer
import cv2 as cv
import numpy as np
from openpyxl import load_workbook
from main.exceptions import InvalidCaptcha, SerialNumberNotAvailable
from main.server import Parser, ResultAPI  
from main.manager import ResultManager


class UploadConsumer(WebsocketConsumer):

    def connect(self):
        self.accept()

    def disconnect(self, close_code):
        pass

    def info(self, type:str, msg:str):
        self.send(text_data=json.dumps({
            "type":type,
            "message":msg
        }))


    def receive(self, text_data=None, bytes_data=None):
 
        data = json.loads(text_data)
        code = data["code"]
        file_data = data["file_data"]
        format, file_str = file_data.split(';base64,')
        file_bytes = base64.b64decode(file_str)
        file = io.BytesIO(file_bytes)
        wb = load_workbook(file)
        sheet = wb.active
        api = ResultAPI(code)
        total_students = sheet.max_row - 1
        current_index = 0
        self.info("total", str(total_students))
        for row in sheet.iter_rows(min_row=2, values_only=True):
            current_index+=1
            self.info("current_index", str(current_index))
            usn = row[0]
            self.info("usn", usn)
            manager = ResultManager(usn, code)
            try:
                student_data, subjects_data  = manager.from_database()
            except:
                while True:
                    try:
                        token = api.get_page()
                        image_bytes = api.get_captcha_image(api.cookies)
                        self.send(bytes_data=image_bytes)
                        nparr = np.frombuffer(image_bytes, np.uint8)
                        img_np = cv.imdecode(nparr, cv.IMREAD_COLOR)
                        processed = api.pre_process(img_np)
                        captcha = api.pred_captcha(processed)
                        self.info("captcha", captcha)
                        student_data, subjects_data = manager.from_vtu(token, captcha, api.cookies)
                        break
                    except InvalidCaptcha as e:
                        self.info("error", f"Invalid captcha: {captcha}, Trying again!")
                        continue
                    except SerialNumberNotAvailable as e:
                        self.info("error", str(e))
                        break
                    except Exception as e:
                        self.info("error", str(e))
                        break
            self.info("success", f"{usn} Done ✅")
            self.info("student", json.dumps({"name":student_data.name, "usn":student_data.usn}))
    
        self.info("success", "Got all student data in the database")

        return super().receive(text_data, bytes_data)
